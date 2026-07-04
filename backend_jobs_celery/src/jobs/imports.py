"""Leitura, staging, deduplicacao e carga aprovada de importacoes."""

import csv
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import load_workbook
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb

from jobs.database import normalize_database_url
from jobs.import_rules import (
    automatic_mapping,
    name_similarity,
    normalize_cpf,
    normalize_date,
    normalize_email,
    normalize_phone,
    split_address,
)


class ImportProcessor:
    def __init__(self, database_url: str) -> None:
        self.database_url = normalize_database_url(database_url)

    def process(self, *, tenant_id: int, import_id: int) -> dict[str, int]:
        with psycopg.connect(self.database_url, row_factory=dict_row) as connection:
            connection.execute(
                "SELECT set_config('app.current_tenant_id', %s, true)",
                (str(tenant_id),),
            )
            config = connection.execute(
                """
                SELECT i.id, i.status, i.mapeamento_colunas, i.parametros,
                       a.caminho, a.extensao
                FROM etl.importacao i
                JOIN etl.importacao_arquivo ia ON ia.importacao_id = i.id
                JOIN arquivo.arquivo a ON a.id = ia.arquivo_id
                WHERE i.tenant_id = %s AND i.id = %s
                """,
                (tenant_id, import_id),
            ).fetchone()
            if not config or config["status"] == "cancelada":
                raise LookupError("Importacao nao encontrada ou cancelada.")
            rows, headers = self._read_rows(
                Path(config["caminho"]),
                str(config["extensao"]),
                dict(config["parametros"] or {}),
            )
            mapping = dict(config["mapeamento_colunas"] or {})
            if not mapping:
                mapping = automatic_mapping(headers)
            if "nome_completo" not in mapping.values():
                raise ValueError("Mapeamento deve possuir o campo nome_completo.")

            connection.execute(
                "UPDATE etl.importacao SET status = 'processando', iniciado_em = now(), "
                "mapeamento_colunas = %s, parametros = parametros || %s "
                "WHERE tenant_id = %s AND id = %s",
                (
                    Jsonb(mapping),
                    Jsonb({"colunas_detectadas": headers}),
                    tenant_id,
                    import_id,
                ),
            )
            connection.execute(
                "DELETE FROM etl.resultado_deduplicacao "
                "WHERE tenant_id = %s AND importacao_id = %s",
                (tenant_id, import_id),
            )
            connection.execute(
                "DELETE FROM etl.staging_pessoa "
                "WHERE tenant_id = %s AND importacao_id = %s",
                (tenant_id, import_id),
            )
            connection.execute(
                "DELETE FROM etl.erro_importacao "
                "WHERE tenant_id = %s AND importacao_id = %s",
                (tenant_id, import_id),
            )
            connection.execute(
                "DELETE FROM etl.importacao_linha "
                "WHERE tenant_id = %s AND importacao_id = %s",
                (tenant_id, import_id),
            )

            valid = invalid = duplicates = 0
            for number, raw in enumerate(rows, start=2):
                inserted_line = connection.execute(
                    """
                    INSERT INTO etl.importacao_linha
                        (tenant_id, importacao_id, numero_linha, conteudo_bruto)
                    VALUES (%s, %s, %s, %s) RETURNING id
                    """,
                    (tenant_id, import_id, number, Jsonb(raw)),
                ).fetchone()
                if inserted_line is None:
                    raise RuntimeError("Falha ao persistir linha da importacao.")
                line_id = int(inserted_line["id"])
                normalized, errors = self._normalize_row(raw, mapping)
                if errors:
                    invalid += 1
                    connection.execute(
                        "UPDATE etl.importacao_linha SET status = 'erro', mensagem = %s "
                        "WHERE id = %s",
                        ("; ".join(message for _, _, message in errors), line_id),
                    )
                    for field, value, message in errors:
                        self._insert_error(
                            connection,
                            tenant_id,
                            import_id,
                            line_id,
                            "validacao",
                            field,
                            value,
                            message,
                        )
                    continue
                staging_id = self._insert_staging(
                    connection, tenant_id, import_id, line_id, normalized
                )
                duplicate = self._find_duplicate(
                    connection,
                    tenant_id,
                    import_id,
                    staging_id,
                    normalized,
                )
                if duplicate:
                    duplicates += 1
                    connection.execute(
                        "UPDATE etl.staging_pessoa SET status = 'duplicado' WHERE id = %s",
                        (staging_id,),
                    )
                    connection.execute(
                        "UPDATE etl.importacao_linha SET status = 'aviso', mensagem = %s "
                        "WHERE id = %s",
                        ("Possivel duplicidade identificada.", line_id),
                    )
                    self._insert_error(
                        connection,
                        tenant_id,
                        import_id,
                        line_id,
                        "deduplicacao",
                        duplicate["criterio"],
                        None,
                        f"Possivel duplicidade ({duplicate['score']:.2f}%).",
                        severity="aviso",
                    )
                else:
                    valid += 1
                    connection.execute(
                        "UPDATE etl.importacao_linha SET status = 'processada' WHERE id = %s",
                        (line_id,),
                    )
            total = valid + invalid + duplicates
            connection.execute(
                """
                UPDATE etl.importacao SET status = 'parcial', total_linhas = %s,
                    linhas_validas = %s, linhas_erro = %s, linhas_duplicadas = %s,
                    linhas_pendentes = %s, linhas_carregadas = 0, concluido_em = now()
                WHERE tenant_id = %s AND id = %s
                """,
                (total, valid, invalid, duplicates, valid + duplicates, tenant_id, import_id),
            )
            return {
                "total": total,
                "validas": valid,
                "invalidas": invalid,
                "duplicadas": duplicates,
            }

    def fail(self, *, tenant_id: int, import_id: int, message: str) -> None:
        with psycopg.connect(self.database_url) as connection:
            connection.execute(
                "SELECT set_config('app.current_tenant_id', %s, true)",
                (str(tenant_id),),
            )
            connection.execute(
                "UPDATE etl.importacao SET status = 'falha', concluido_em = now() "
                "WHERE tenant_id = %s AND id = %s AND status <> 'cancelada'",
                (tenant_id, import_id),
            )
            connection.execute(
                """
                INSERT INTO etl.erro_importacao
                    (tenant_id, importacao_id, etapa, mensagem, severidade)
                VALUES (%s, %s, 'leitura', %s, 'erro')
                """,
                (tenant_id, import_id, message[:1000]),
            )

    def load(self, *, tenant_id: int, import_id: int) -> dict[str, int]:
        loaded = failed = 0
        with psycopg.connect(self.database_url, row_factory=dict_row) as connection:
            connection.execute(
                "SELECT set_config('app.current_tenant_id', %s, true)",
                (str(tenant_id),),
            )
            config = connection.execute(
                "SELECT fonte_dado_id, aprovado_em, status FROM etl.importacao "
                "WHERE tenant_id = %s AND id = %s",
                (tenant_id, import_id),
            ).fetchone()
            if not config or not config["aprovado_em"] or config["status"] == "cancelada":
                raise ValueError("Importacao nao aprovada para carga.")
            staging_rows = connection.execute(
                """
                SELECT * FROM etl.staging_pessoa
                WHERE tenant_id = %s AND importacao_id = %s AND status = 'validado'
                ORDER BY id
                """,
                (tenant_id, import_id),
            ).fetchall()
            for row in staging_rows:
                try:
                    with connection.transaction():
                        person_id = self._load_person(
                            connection,
                            tenant_id,
                            int(config["fonte_dado_id"]),
                            dict(row),
                        )
                        connection.execute(
                            "UPDATE etl.staging_pessoa SET status = 'carregado', pessoa_id = %s "
                            "WHERE id = %s",
                            (person_id, row["id"]),
                        )
                        connection.execute(
                            "UPDATE etl.importacao_linha SET status = 'processada' "
                            "WHERE id = %s",
                            (row["importacao_linha_id"],),
                        )
                    loaded += 1
                except Exception as exc:
                    failed += 1
                    connection.execute(
                        "UPDATE etl.importacao_linha SET status = 'erro', mensagem = %s "
                        "WHERE id = %s",
                        (str(exc)[:500], row["importacao_linha_id"]),
                    )
                    self._insert_error(
                        connection,
                        tenant_id,
                        import_id,
                        int(row["importacao_linha_id"]),
                        "carga",
                        None,
                        None,
                        "Falha ao promover registro para cadastro.",
                    )
            connection.execute(
                """
                UPDATE etl.importacao SET status = %s, linhas_carregadas = %s,
                    linhas_erro = COALESCE(linhas_erro, 0) + %s,
                    linhas_pendentes = GREATEST(COALESCE(linhas_pendentes, 0) - %s, 0),
                    concluido_em = now()
                WHERE tenant_id = %s AND id = %s
                """,
                (
                    "concluida" if failed == 0 else "parcial",
                    loaded,
                    failed,
                    loaded,
                    tenant_id,
                    import_id,
                ),
            )
        return {"carregadas": loaded, "falhas": failed}

    @staticmethod
    def _read_rows(
        path: Path, extension: str, parameters: dict[str, Any]
    ) -> tuple[list[dict[str, Any]], list[str]]:
        if extension == "csv":
            content: str
            try:
                content = path.read_text(encoding=str(parameters.get("encoding", "utf-8-sig")))
            except UnicodeDecodeError:
                content = path.read_text(encoding="latin-1")
            sample = content[:4096]
            delimiter = str(parameters.get("separador") or "")
            if not delimiter:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
            reader = csv.DictReader(content.splitlines(), delimiter=delimiter)
            headers = [str(header) for header in (reader.fieldnames or [])]
            return (
                [
                    {str(key): value for key, value in row.items()}
                    for row in reader
                    if any(value not in (None, "") for value in row.values())
                ],
                headers,
            )
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = parameters.get("aba")
        sheet = workbook[str(sheet_name)] if sheet_name else workbook.worksheets[0]
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(values)]
        rows = [
            {headers[index]: value for index, value in enumerate(row)}
            for row in values
            if any(value not in (None, "") for value in row)
        ]
        workbook.close()
        return rows, headers

    @staticmethod
    def _normalize_row(
        raw: dict[str, Any], mapping: dict[str, str]
    ) -> tuple[dict[str, Any], list[tuple[str, str | None, str]]]:
        data = {
            target: raw.get(source)
            for source, target in mapping.items()
            if target
        }
        errors: list[tuple[str, str | None, str]] = []
        name = str(data.get("nome_completo") or "").strip()
        if not name:
            errors.append(("nome_completo", None, "Nome completo e obrigatorio."))
        normalized: dict[str, Any] = {"nome_completo": name}
        functions = {
            "cpf": normalize_cpf,
            "telefone": normalize_phone,
            "email": normalize_email,
            "data_nascimento": normalize_date,
        }
        for field, function in functions.items():
            try:
                normalized[field] = function(data.get(field))
            except ValueError as exc:
                errors.append((field, str(data.get(field) or "")[:200], str(exc)))
        normalized["rg"] = str(data.get("rg") or "").strip() or None
        normalized["titulo_eleitor"] = (
            "".join(
                character
                for character in str(data.get("titulo_eleitor") or "")
                if character.isdigit()
            )
            or None
        )
        try:
            normalized.update(split_address(data))
        except ValueError as exc:
            errors.append(("cep", str(data.get("cep") or "")[:200], str(exc)))
        normalized["dados_extras"] = {
            key: value
            for key, value in raw.items()
            if key not in mapping or not mapping[key]
        }
        return normalized, errors

    @staticmethod
    def _insert_staging(
        connection: psycopg.Connection[Any],
        tenant_id: int,
        import_id: int,
        line_id: int,
        data: dict[str, Any],
    ) -> int:
        inserted = connection.execute(
            """
            INSERT INTO etl.staging_pessoa
                (tenant_id, importacao_id, importacao_linha_id, nome_completo,
                 cpf, rg, titulo_eleitor, data_nascimento, telefone, email,
                 endereco, logradouro, numero, complemento, bairro, municipio,
                 uf, cep, dados_extras, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, 'validado')
            RETURNING id
            """,
            (
                tenant_id,
                import_id,
                line_id,
                data["nome_completo"],
                data.get("cpf"),
                data.get("rg"),
                data.get("titulo_eleitor"),
                data.get("data_nascimento"),
                data.get("telefone"),
                data.get("email"),
                data.get("endereco"),
                data.get("logradouro"),
                data.get("numero"),
                data.get("complemento"),
                data.get("bairro"),
                data.get("municipio"),
                data.get("uf"),
                data.get("cep"),
                Jsonb(data.get("dados_extras", {})),
            ),
        ).fetchone()
        if inserted is None:
            raise RuntimeError("Falha ao persistir staging.")
        return int(inserted["id"])

    def _find_duplicate(
        self,
        connection: psycopg.Connection[Any],
        tenant_id: int,
        import_id: int,
        staging_id: int,
        data: dict[str, Any],
    ) -> dict[str, Any] | None:
        checks: list[tuple[str, str, Any]] = []
        if data.get("cpf"):
            checks.append(
                (
                    "cpf",
                    "SELECT p.id FROM cadastro.pessoa p JOIN cadastro.pessoa_documento d "
                    "ON d.pessoa_id = p.id WHERE p.tenant_id = %s "
                    "AND d.tipo_documento = 'cpf' AND d.numero = %s LIMIT 1",
                    data["cpf"],
                )
            )
        if data.get("titulo_eleitor"):
            checks.append(
                (
                    "titulo_eleitor",
                    "SELECT p.id FROM cadastro.pessoa p JOIN cadastro.pessoa_documento d "
                    "ON d.pessoa_id = p.id WHERE p.tenant_id = %s "
                    "AND d.tipo_documento = 'titulo_eleitor' AND d.numero = %s LIMIT 1",
                    data["titulo_eleitor"],
                )
            )
        for contact_criterion, contact_type, value in (
            ("telefone", "telefone", data.get("telefone")),
            ("email", "email", data.get("email")),
        ):
            if value:
                checks.append(
                    (
                        contact_criterion,
                        "SELECT p.id FROM cadastro.pessoa p JOIN cadastro.pessoa_contato c "
                        "ON c.pessoa_id = p.id WHERE p.tenant_id = %s "
                        "AND c.tipo_contato = %s AND c.valor = %s LIMIT 1",
                        (contact_type, value),
                    )
                )
        candidate_id: int | None = None
        criterion: str | None = None
        for name, query, value in checks:
            params = (
                (tenant_id, *value)
                if isinstance(value, tuple)
                else (tenant_id, value)
            )
            row = connection.execute(query, params).fetchone()
            if row:
                candidate_id, criterion = int(row["id"]), name
                break
        score = 100.0
        other_staging_id: int | None = None
        if candidate_id is None:
            exact = connection.execute(
                """
                SELECT id FROM etl.staging_pessoa
                WHERE tenant_id = %s AND importacao_id = %s AND id <> %s
                  AND status IN ('validado', 'duplicado')
                  AND ((%s::text IS NOT NULL AND cpf = %s::text)
                    OR (%s::text IS NOT NULL AND telefone = %s::text)
                    OR (%s::text IS NOT NULL AND email = %s::text))
                ORDER BY id LIMIT 1
                """,
                (
                    tenant_id,
                    import_id,
                    staging_id,
                    data.get("cpf"),
                    data.get("cpf"),
                    data.get("telefone"),
                    data.get("telefone"),
                    data.get("email"),
                    data.get("email"),
                ),
            ).fetchone()
            if exact:
                other_staging_id, criterion = int(exact["id"]), "arquivo_interno"
        if candidate_id is None and other_staging_id is None:
            candidates = connection.execute(
                "SELECT id, nome_completo, data_nascimento FROM cadastro.pessoa "
                "WHERE tenant_id = %s AND ativo AND excluido_em IS NULL LIMIT 500",
                (tenant_id,),
            ).fetchall()
            best: tuple[int, float] | None = None
            for candidate in candidates:
                similarity = name_similarity(
                    data["nome_completo"], candidate["nome_completo"]
                )
                same_birth = (
                    data.get("data_nascimento")
                    and candidate["data_nascimento"] == data["data_nascimento"]
                )
                if similarity >= 85 and (same_birth or similarity >= 92):
                    if best is None or similarity > best[1]:
                        best = (int(candidate["id"]), similarity)
            if best:
                candidate_id, score, criterion = best[0], best[1], "fuzzy"
        if criterion is None:
            return None
        rule_id = connection.execute(
            "SELECT id FROM etl.regra_deduplicacao "
            "WHERE (tenant_id IS NULL OR tenant_id = %s) AND criterio = %s "
            "AND ativa ORDER BY tenant_id NULLS FIRST LIMIT 1",
            (tenant_id, criterion if criterion != "arquivo_interno" else "cpf"),
        ).fetchone()
        connection.execute(
            """
            INSERT INTO etl.resultado_deduplicacao
                (tenant_id, regra_deduplicacao_id, importacao_id,
                 staging_pessoa_id, pessoa_candidata_id,
                 registro_origem_id, registro_duplicado_id,
                 score, decisao, detalhes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pendente', %s)
            """,
            (
                tenant_id,
                rule_id["id"] if rule_id else None,
                import_id,
                staging_id,
                candidate_id,
                staging_id,
                candidate_id or other_staging_id,
                score,
                Jsonb({"criterio": criterion}),
            ),
        )
        return {"criterio": criterion, "score": score}

    @staticmethod
    def _insert_error(
        connection: psycopg.Connection[Any],
        tenant_id: int,
        import_id: int,
        line_id: int,
        stage: str,
        field: str | None,
        value: str | None,
        message: str,
        *,
        severity: str = "erro",
    ) -> None:
        connection.execute(
            """
            INSERT INTO etl.erro_importacao
                (tenant_id, importacao_id, importacao_linha_id, etapa,
                 campo, valor, mensagem, severidade)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (tenant_id, import_id, line_id, stage, field, value, message, severity),
        )

    @staticmethod
    def _load_person(
        connection: psycopg.Connection[Any],
        tenant_id: int,
        source_id: int,
        row: dict[str, Any],
    ) -> int:
        inserted_person = connection.execute(
            """
            INSERT INTO cadastro.pessoa
                (tenant_id, nome_completo, data_nascimento, fonte_dado_id)
            VALUES (%s, %s, %s, %s) RETURNING id
            """,
            (
                tenant_id,
                row["nome_completo"],
                row["data_nascimento"],
                source_id,
            ),
        ).fetchone()
        if inserted_person is None:
            raise RuntimeError("Falha ao criar pessoa.")
        person_id = int(inserted_person["id"])
        for document_type, value in (
            ("cpf", row["cpf"]),
            ("rg", row["rg"]),
            ("titulo_eleitor", row["titulo_eleitor"]),
        ):
            if value:
                connection.execute(
                    "INSERT INTO cadastro.pessoa_documento "
                    "(tenant_id, pessoa_id, tipo_documento, numero) "
                    "VALUES (%s, %s, %s, %s)",
                    (tenant_id, person_id, document_type, value),
                )
        for contact_type, value in (
            ("telefone", row["telefone"]),
            ("email", row["email"]),
        ):
            if value:
                connection.execute(
                    "INSERT INTO cadastro.pessoa_contato "
                    "(tenant_id, pessoa_id, tipo_contato, valor, principal) "
                    "VALUES (%s, %s, %s, %s, TRUE)",
                    (tenant_id, person_id, contact_type, value),
                )
        if row["titulo_eleitor"]:
            connection.execute(
                "INSERT INTO cadastro.eleitor (tenant_id, pessoa_id, titulo_eleitor) "
                "VALUES (%s, %s, %s)",
                (tenant_id, person_id, row["titulo_eleitor"]),
            )
        if any(row.get(field) for field in ("logradouro", "bairro", "cep")):
            municipality_id = None
            if row.get("municipio"):
                municipality = connection.execute(
                    """
                    SELECT m.id FROM global.municipio m
                    JOIN global.estado e ON e.id = m.estado_id
                    WHERE unaccent(lower(m.nome)) = unaccent(lower(%s))
                      AND (%s::text IS NULL OR e.uf = %s::text)
                    LIMIT 1
                    """,
                    (row["municipio"], row.get("uf"), row.get("uf")),
                ).fetchone()
                municipality_id = municipality["id"] if municipality else None
            inserted_address = connection.execute(
                """
                INSERT INTO cadastro.endereco
                    (tenant_id, municipio_id, bairro_texto, logradouro,
                     numero, complemento, cep)
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
                """,
                (
                    tenant_id,
                    municipality_id,
                    row["bairro"],
                    row["logradouro"],
                    row["numero"],
                    row["complemento"],
                    row["cep"],
                ),
            ).fetchone()
            if inserted_address is None:
                raise RuntimeError("Falha ao criar endereco.")
            address_id = int(inserted_address["id"])
            connection.execute(
                "INSERT INTO cadastro.pessoa_endereco "
                "(tenant_id, pessoa_id, endereco_id, principal) "
                "VALUES (%s, %s, %s, TRUE)",
                (tenant_id, person_id, address_id),
            )
        return person_id
